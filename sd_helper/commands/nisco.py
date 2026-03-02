from __future__ import annotations

import io
import json
from pathlib import Path

import click
import httpx


NISCO_BASE_URL = "http://172.25.45.183:5001"
DEFAULT_ROWS = 50
IMAGE_SIZE = 512       # Embedded image size in Excel (pixels)
IMAGE_COL_WIDTH = 40   # Excel column width in chars


@click.group()
def nisco():
    """Nisco system commands."""
    pass


@nisco.group()
def ocr():
    """OCR-related commands for Nisco."""
    pass


@ocr.command("image-download")
@click.option("--start", "start_date", required=True, metavar="YYYY-MM-DD",
              help="Start date (inclusive), filters date_time field")
@click.option("--end", "end_date", required=True, metavar="YYYY-MM-DD",
              help="End date (inclusive), filters date_time field")
@click.option("--output", "-o", default="nisco_ocr.xlsx", show_default=True,
              help="Output Excel file path")
@click.option("--rows", "-r", default=DEFAULT_ROWS, show_default=True, type=int,
              help="Rows per page when fetching")
@click.option("--images-dir", "-i", default=None,
              help="Directory to save original images (default: <output_stem>_images/)")
@click.option("--no-verify", is_flag=True, help="Disable SSL certificate verification")
@click.option("--debug", is_flag=True, help="Print debug info")
def image_download(start_date, end_date, output, rows, images_dir, no_verify, debug):
    """Download Nisco OCR records with images into an Excel file.

    Fetches all pages for the given date range, saves original images to disk,
    and embeds them into the output Excel spreadsheet.

    \b
    Examples:
      sd-helper nisco ocr image-download --start 2026-03-01 --end 2026-03-02
      sd-helper nisco ocr image-download --start 2026-03-01 --end 2026-03-31 -o march.xlsx
      sd-helper nisco ocr image-download --start 2026-03-01 --end 2026-03-02 -i ./photos
    """
    try:
        import openpyxl
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment
    except ImportError:
        raise click.ClickException("openpyxl is required. Run: pip install openpyxl")

    verify = not no_verify
    out_path = Path(output)
    img_dir = Path(images_dir) if images_dir else out_path.parent / f"{out_path.stem}_images"
    img_dir.mkdir(parents=True, exist_ok=True)

    click.echo(f"Fetching records from {start_date} to {end_date}...")
    records = _fetch_all_pages(start_date, end_date, rows, verify, debug)
    if not records:
        click.echo("No records found.")
        return

    click.echo(f"Found {len(records)} records, filtering result_sta=否...")
    records = [r for r in records if r.get("result_sta") == "否"]
    click.echo(f"{len(records)} records after filter. Downloading images and building Excel...")
    if not records:
        click.echo("No records matched the filter.")
        return
    click.echo(f"Saving images to: {img_dir.resolve()}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OCR Records"

    # Determine column order: all keys from first record, then "image_file_name", then "image"
    field_keys = list(records[0].keys())
    headers = field_keys + ["image_file_name", "image"]
    img_name_col_idx = len(headers) - 1  # 1-based index of image_file_name column
    img_col_idx = len(headers)           # 1-based index of image column

    # Write header row
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    img_cache: dict[str, bytes | None] = {}
    with httpx.Client(verify=verify, timeout=15) as client:
        for row_idx, record in enumerate(records, start=2):
            ws.row_dimensions[row_idx].height = IMAGE_SIZE * 0.75  # pt = px * 0.75

            for col, key in enumerate(field_keys, start=1):
                cell = ws.cell(row=row_idx, column=col, value=_safe_value(record.get(key)))
                cell.alignment = Alignment(wrap_text=False, vertical="center")

            pic_url = record.get("pic_address", "")
            img_bytes = _fetch_image(client, pic_url, img_cache, debug)

            if img_bytes:
                # Save original image to disk
                filename = _filename_from_url(pic_url)
                img_path = img_dir / filename
                img_path.write_bytes(img_bytes)
                ws.cell(row=row_idx, column=img_name_col_idx, value=filename)

                try:
                    img = XLImage(io.BytesIO(img_bytes))
                    scale = IMAGE_SIZE / max(img.height, img.width, 1)
                    img.width = int(img.width * scale)
                    img.height = int(img.height * scale)
                    cell_ref = f"{get_column_letter(img_col_idx)}{row_idx}"
                    ws.add_image(img, cell_ref)
                except Exception as e:
                    if debug:
                        click.echo(f"[DEBUG] Image embed failed row {row_idx}: {e}", err=True)
                    ws.cell(row=row_idx, column=img_col_idx, value=pic_url)
            else:
                ws.cell(row=row_idx, column=img_col_idx, value=pic_url)

            if row_idx % 50 == 0:
                click.echo(f"  Processed {row_idx - 1}/{len(records)} records...")

    # Set image column width
    ws.column_dimensions[get_column_letter(img_col_idx)].width = IMAGE_COL_WIDTH
    ws.row_dimensions[1].height = 20  # normal header row

    out_path = Path(output)
    wb.save(out_path)
    click.echo(f"\nSaved {len(records)} records to {out_path.resolve()}")


NISCO_OCR_PROMPT = (
    "这是一张工业标签图片，标签上有手写数字。"
    "请识别图片中的所有数字内容，只返回JSON格式。"
    "格式：{\"result\": \"识别的内容\", \"reason\": \"说明\"}"
)


@ocr.command("run")
@click.argument("excel_file", type=click.Path(exists=True))
@click.option("--images-dir", "-i", default=None,
              help="Directory containing images (default: <excel_stem>_images/)")
@click.option("--model", "-m", default=None, help="Model name (from config)")
@click.option("--profile", default=None, help="IAM profile for authentication")
@click.option("--prompt", "-p", default=None, help="Custom prompt text")
@click.option("--prompt-file", type=click.Path(exists=True), default=None,
              help="File containing the prompt text")
@click.option("--no-verify", is_flag=True, help="Disable SSL certificate verification")
@click.option("--debug", is_flag=True, help="Print debug info")
def ocr_run(excel_file, images_dir, model, profile, prompt, prompt_file, no_verify, debug):
    """Run vision OCR on images from a downloaded Excel file.

    Reads image_file_name from the Excel, calls a multimodal LLM for each image,
    and writes the results back into the same Excel file as new columns.

    \b
    Examples:
      sd-helper nisco ocr run nisco_ocr.xlsx
      sd-helper nisco ocr run nisco_ocr.xlsx -m qwen3-vl-32b
      sd-helper nisco ocr run nisco_ocr.xlsx --prompt "识别图片中的炉号"
      sd-helper nisco ocr run nisco_ocr.xlsx --prompt-file my_prompt.txt
    """
    from ..auth import get_token_from_config, load_config
    from ..api import LLMClient, build_vision_message, bytes_to_data_url, get_model_config

    try:
        import openpyxl
    except ImportError:
        raise click.ClickException("openpyxl is required. Run: pip install openpyxl")

    try:
        from PIL import Image as PILImage
    except ImportError:
        raise click.ClickException("Pillow is required. Run: pip install Pillow")

    excel_path = Path(excel_file)
    img_dir = Path(images_dir) if images_dir else excel_path.parent / f"{excel_path.stem}_images"
    if not img_dir.exists():
        raise click.ClickException(f"Images directory not found: {img_dir}")

    # Determine prompt
    if prompt_file:
        effective_prompt = Path(prompt_file).read_text(encoding="utf-8").strip()
    elif prompt:
        effective_prompt = prompt
    else:
        effective_prompt = NISCO_OCR_PROMPT

    if debug:
        click.echo(f"[DEBUG] Prompt: {effective_prompt[:100]}...", err=True)

    # Set up LLM client
    config = load_config(profile)
    model_config = get_model_config(config, model)
    if not model_config:
        raise click.ClickException("No vision model configured. Use --model or set default_model in config.")

    try:
        token_info = get_token_from_config(profile=profile)
        token = token_info["token"]
    except ValueError as e:
        raise click.ClickException(str(e))

    effective_verify = False if no_verify else model_config.verify_ssl
    llm_client = LLMClient(
        endpoint=model_config.endpoint,
        token=token,
        model_type=model_config.type,
        verify_ssl=effective_verify,
    )

    # Load Excel
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Find image_file_name column and last column
    header_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if "image_file_name" not in header_row:
        raise click.ClickException("'image_file_name' column not found in Excel. Run image-download first.")

    img_name_col = header_row.index("image_file_name") + 1  # 1-based

    # Add result columns after existing columns (idempotent: reuse if already present)
    result_cols = ["ocr_raw", "ocr_result", "ocr_reason"]
    col_map = {}
    for name in result_cols:
        if name in header_row:
            col_map[name] = header_row.index(name) + 1
        else:
            next_col = ws.max_column + 1
            ws.cell(1, next_col, value=name)
            col_map[name] = next_col

    total_rows = ws.max_row - 1  # exclude header
    click.echo(f"Processing {total_rows} rows from {excel_path.name}...")
    click.echo(f"Images dir: {img_dir.resolve()}")

    processed = skipped = errors = 0
    for row_idx in range(2, ws.max_row + 1):
        filename = ws.cell(row_idx, img_name_col).value
        if not filename:
            skipped += 1
            continue

        img_path = img_dir / filename
        if not img_path.exists():
            click.echo(f"  [MISSING] {filename}", err=True)
            ws.cell(row_idx, col_map["ocr_raw"], value="image not found")
            errors += 1
            continue

        try:
            pil_img = PILImage.open(img_path)
            buf = io.BytesIO()
            pil_img.save(buf, format="JPEG", quality=95)
            url = bytes_to_data_url(buf.getvalue(), "image/jpeg")
            messages = [build_vision_message(effective_prompt, [url])]
        except Exception as e:
            click.echo(f"  [ERROR] {filename}: {e}", err=True)
            ws.cell(row_idx, col_map["ocr_raw"], value=f"load error: {e}")
            errors += 1
            continue

        # Call LLM
        from .llm import _send_chat
        response = _send_chat(llm_client, messages, temperature=0.1, max_tokens=256,
                              stream=False, debug=debug, silent=True)

        if response:
            response = response.strip().removeprefix("```json").removeprefix("```").removesuffix("```").strip()

        ws.cell(row_idx, col_map["ocr_raw"], value=_safe_value(response or ""))

        try:
            data = json.loads(response) if response else {}
        except json.JSONDecodeError:
            data = {"result": response, "reason": ""}

        ws.cell(row_idx, col_map["ocr_result"], value=_safe_value(str(data.get("result", ""))))
        ws.cell(row_idx, col_map["ocr_reason"], value=_safe_value(str(data.get("reason", ""))))

        processed += 1
        click.echo(f"  [{processed}/{total_rows}] {filename}: {data.get('result', '')}")

    wb.save(excel_path)
    click.echo(f"\nDone. {processed} processed, {skipped} skipped, {errors} errors.")
    click.echo(f"Results written to {excel_path.resolve()}")


def _build_params(page: int, rows: int, search_time: str) -> dict:
    return {
        "page": page,
        "rows_per_page": rows,
        "search_term": "",
        "column_index": 0,
        "search_term2": "",
        "column_index2": 0,
        "search_term3": "",
        "column_index3": 0,
        "search_time": search_time,
    }



def _in_range(record: dict, start_date: str, end_date: str) -> bool:
    dt = record.get("date_time", "")
    if not dt:
        return True  # include records with no date
    return start_date <= dt[:10] <= end_date


def _fetch_all_pages(start_date: str, end_date: str, rows: int, verify: bool, debug: bool) -> list[dict]:
    url = f"{NISCO_BASE_URL}/get_data"
    all_records: list[dict] = []
    page = 1

    with httpx.Client(verify=verify, timeout=30) as client:
        while True:
            params = _build_params(page, rows, "%")
            if debug:
                click.echo(f"[DEBUG] GET {url} page={page}", err=True)

            try:
                resp = client.get(url, params=params)
                resp.raise_for_status()
            except httpx.HTTPError as e:
                raise click.ClickException(f"HTTP error on page {page}: {e}")

            body = resp.json()
            page_records = _extract_records(body)

            if not page_records:
                break

            # Client-side date filter
            filtered = [r for r in page_records if _in_range(r, start_date, end_date)]
            all_records.extend(filtered)

            if debug:
                click.echo(f"[DEBUG] Page {page}: {len(page_records)} records, {len(filtered)} in range", err=True)

            if len(page_records) < rows:
                break  # last page

            # API returns newest-first: if this whole page is before start_date, we're done
            dates = [r.get("date_time", "")[:10] for r in page_records if r.get("date_time")]
            if dates and all(d < start_date for d in dates):
                break

            page += 1

    return all_records


def _extract_records(body) -> list[dict]:
    if isinstance(body, list):
        return body
    for key in ("data", "rows", "records", "items", "results"):
        if key in body and isinstance(body[key], list):
            return body[key]
    return []


def _filename_from_url(url: str) -> str:
    """Extract filename from URL, fallback to URL hash."""
    from urllib.parse import urlparse, unquote
    path = urlparse(url).path
    name = unquote(path.split("/")[-1]) if path else ""
    return name if name else f"{hash(url) & 0xFFFFFF}.jpg"


def _safe_value(value):
    """Strip control characters that openpyxl rejects."""
    import re
    if isinstance(value, str):
        return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", value)
    return value


def _fetch_image(client: httpx.Client, url: str, cache: dict, debug: bool) -> bytes | None:
    if not url:
        return None
    if url in cache:
        return cache[url]
    try:
        resp = client.get(url)
        resp.raise_for_status()
        data = resp.content
    except Exception as e:
        if debug:
            click.echo(f"[DEBUG] Failed to fetch image {url}: {e}", err=True)
        data = None
    cache[url] = data
    return data
